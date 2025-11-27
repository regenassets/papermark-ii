#!/usr/bin/env python3
"""
Verify the DYNAMIC_DROPDOWNS model has correct data validation and formulas
"""

import openpyxl
from pathlib import Path

FILE = Path('/home/cyrus/cyrus-app/concepts/RealGold_Finmodel_V2_DYNAMIC_DROPDOWNS.xlsx')

def print_section(title):
    print(f"\n{'='*80}")
    print(f"{title}")
    print('='*80)

def main():
    print_section("Verifying DYNAMIC_DROPDOWNS Model")

    wb = openpyxl.load_workbook(FILE)

    # Check XAUconfig timeline
    print_section("XAUconfig Timeline")
    config_ws = wb['XAUconfig']
    print(f"B4 (First month label): {config_ws['B4'].value}")
    print(f"B6 (First dropdown item): {config_ws['B6'].value}")
    print(f"BI4 (60th month label): {config_ws['BI4'].value}")
    print(f"BI6 (60th dropdown item): {config_ws['BI6'].value}")

    # Check Mine Inventory
    print_section("Mine Inventory")
    mine_ws = wb['Mine Inventory']
    print(f"B2 (Courbet date): {mine_ws['B2'].value}")
    print(f"B3 (Mine 2 date): {mine_ws['B3'].value}")
    print(f"\nAO1 (Header): {mine_ws['AO1'].value}")
    print(f"AO2 (Courbet offset formula): {mine_ws['AO2'].value}")
    print(f"AO3 (Mine 2 offset formula): {mine_ws['AO3'].value}")

    # Check data validation
    print("\nData Validation (Dropdowns):")
    dv_count = 0
    for dv in mine_ws.data_validations.dataValidation:
        print(f"  Type: {dv.type}")
        print(f"  Formula1 (source): {dv.formula1}")
        print(f"  Show dropdown: {dv.showDropDown}")
        print(f"  Applies to cells: {dv.sqref}")
        dv_count += 1

    if dv_count > 0:
        print(f"\n✓ Found {dv_count} data validation rule(s)")
    else:
        print("\n✗ No data validation found!")

    # Check Resource Supply
    print_section("Resource Supply (5yr)")
    resource_ws = wb['Resource Supply (5yr)']
    print(f"B1 (First header): {resource_ws['B1'].value}")
    print(f"C1 (Second header): {resource_ws['C1'].value}")
    print(f"\nB3 (Mine count month 0): {resource_ws['B3'].value}")
    print(f"C3 (Mine count month 1): {resource_ws['C3'].value}")
    print(f"\nB5 (Gold production month 0): {resource_ws['B5'].value}")
    print(f"C5 (Gold production month 1): {resource_ws['C5'].value}")

    # Check other sheets
    print_section("Other Time-Based Sheets")
    sheets_to_check = [
        'Token Supply (5yr)',
        'Token Demand (5yr)',
        'RA LLC Cashflow',
        'RAF Cashflow (5yr)',
        'RGT Cashflow (5yr)',
    ]

    for sheet_name in sheets_to_check:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            b1_formula = ws['B1'].value
            print(f"  {sheet_name}: B1 = {b1_formula}")

    # Check Inputs
    print_section("Inputs Sheet")
    inputs_ws = wb['Inputs']
    print(f"B26 (Unitization fee): {inputs_ws['B26'].value}")
    print(f"B25 (Admin fee): {inputs_ws['B25'].value}")

    # Summary
    print_section("VERIFICATION SUMMARY")

    checks = []

    # Timeline
    if config_ws['B4'].value == "Dec '25" and config_ws['B6'].value == "Dec '25":
        checks.append("✓ Timeline and dropdown list both start at Dec '25")
    else:
        checks.append(f"✗ Timeline/dropdown mismatch: B4={config_ws['B4'].value}, B6={config_ws['B6'].value}")

    # Dates
    if mine_ws['B2'].value == "Dec '25":
        checks.append("✓ Courbet date is Dec '25")
    else:
        checks.append(f"✗ Courbet date is {mine_ws['B2'].value}, expected Dec '25")

    # Data validation
    if dv_count > 0:
        checks.append(f"✓ Data validation configured ({dv_count} rule(s))")
    else:
        checks.append("✗ No data validation found")

    # MATCH formulas
    if mine_ws['AO2'].value and 'MATCH' in str(mine_ws['AO2'].value):
        checks.append("✓ AO2 contains MATCH formula")
    else:
        checks.append(f"✗ AO2 is not a MATCH formula: {mine_ws['AO2'].value}")

    # SUMIF formulas
    if resource_ws['B3'].value and 'COUNTIF' in str(resource_ws['B3'].value):
        checks.append("✓ Resource Supply B3 contains COUNTIF formula")
    else:
        checks.append(f"✗ Resource Supply B3: {resource_ws['B3'].value}")

    if resource_ws['B5'].value and 'SUMIF' in str(resource_ws['B5'].value):
        checks.append("✓ Resource Supply B5 contains SUMIF formula")
    else:
        checks.append(f"✗ Resource Supply B5: {resource_ws['B5'].value}")

    # Fees
    if inputs_ws['B26'].value == 0.0001:
        checks.append("✓ Unitization fee is 0.0001 (0.01%)")
    else:
        checks.append(f"✗ Unitization fee is {inputs_ws['B26'].value}, expected 0.0001")

    # Other sheet headers
    all_linked = True
    for sheet_name in sheets_to_check:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if not (ws['B1'].value and 'XAUconfig' in str(ws['B1'].value)):
                all_linked = False
                break

    if all_linked:
        checks.append("✓ All time-based sheets linked to XAUconfig")
    else:
        checks.append("✗ Some sheets not linked to XAUconfig")

    for check in checks:
        print(check)

    if all('✓' in c for c in checks):
        print("\n🎉 ALL CHECKS PASSED - Dropdowns configured correctly!")
        return 0
    else:
        print("\n⚠ Some checks failed - see above")
        return 1

if __name__ == '__main__':
    import sys
    sys.exit(main())
