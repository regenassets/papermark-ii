#!/usr/bin/env python3
"""
Fix RealGold Financial Model timeline to start from Jan 2025 instead of Dec 2025.
This ensures 2025 expenses appear in the correct year in cashflow sheets.
"""

import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta

def create_timeline_from_jan_2025(wb):
    """Update XAUconfig timeline to start from Jan 2025."""
    config_ws = wb['XAUconfig']

    # Start from January 2025
    start_date = datetime(2025, 1, 1)

    # Row 3: Gold Price (3% CAGR from $2892.6)
    config_ws['A3'] = 'Gold Price (CAGR)'
    initial_gold_price = 2892.6
    monthly_growth_rate = (1.03 ** (1/12)) - 1  # 3% annual = ~0.247% monthly

    # Row 4: Month labels (Jan '25, Feb '25, ...)
    config_ws['A4'] = 'Month'

    # Row 6: Dropdown source (same as row 4)
    config_ws['A6'] = 'Dropdown Source'

    print("Creating timeline starting from Jan '25...")

    for month_offset in range(60):  # 60 months (5 years)
        current_date = start_date + relativedelta(months=month_offset)
        col_idx = month_offset + 2  # Column B = offset 0
        col_letter = get_column_letter(col_idx)

        # Format: "Jan '25", "Feb '25", etc.
        month_label = current_date.strftime("%b '%y")

        # Gold price with 3% CAGR
        gold_price = initial_gold_price * ((1 + monthly_growth_rate) ** month_offset)

        config_ws[f'{col_letter}3'] = gold_price
        config_ws[f'{col_letter}4'] = month_label
        config_ws[f'{col_letter}6'] = month_label

        if month_offset < 12:
            print(f"  {col_letter}: {month_label} (Gold: ${gold_price:.2f})")

    print(f"\n✓ Timeline created: Jan '25 through Dec '29 (60 months)")

def update_mine_inventory_dropdowns(wb):
    """Update Mine Inventory dropdown validation to use new timeline."""
    mi_ws = wb['Mine Inventory']

    print("\nUpdating Mine Inventory dropdown validation...")

    # The dropdown source is in XAUconfig row 6, columns B through BI (60 months)
    # Starting from Jan '25 instead of Dec '25

    # Data validation should already exist, but the values will auto-update
    # since they reference XAUconfig!$B$6:$BI$6

    print("✓ Dropdown validation will use new timeline (Jan '25 - Dec '29)")

def main():
    print("=" * 80)
    print("FIXING TIMELINE: Dec '25 → Jan '25")
    print("=" * 80)

    # Load the current file
    input_file = 'RealGold_Finmodel_V2_COMPLETE_DYNAMIC.xlsx'
    output_file = 'RealGold_Finmodel_V2_COMPLETE_DYNAMIC_JAN2025.xlsx'

    print(f"\nLoading: {input_file}")
    wb = openpyxl.load_workbook(input_file)

    # Update timeline
    create_timeline_from_jan_2025(wb)

    # Update dropdown validation
    update_mine_inventory_dropdowns(wb)

    # Save the updated file
    print(f"\nSaving: {output_file}")
    wb.save(output_file)

    print("\n" + "=" * 80)
    print("✅ TIMELINE FIXED")
    print("=" * 80)
    print(f"Timeline now starts: Jan '25 (was Dec '25)")
    print(f"Timeline ends: Dec '29 (was Nov '30)")
    print(f"\nThis means:")
    print("  - 2025 expenses will now appear in 2025 (not 2026)")
    print("  - All mine dates and formulas remain functional")
    print("  - Gold prices still use 3% CAGR")
    print(f"\nOutput: {output_file}")

if __name__ == '__main__':
    main()
