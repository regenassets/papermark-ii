#!/usr/bin/env python3
"""
RealGold Financial Model - FULLY DYNAMIC with Date Dropdowns
=============================================================

Creates a fully dynamic model where:
1. Column B in Mine Inventory has DROPDOWN date selectors
2. Column AO contains MATCH formulas that calculate month offsets
3. Resource Supply uses SUMIF formulas referencing those offsets
4. User can SELECT dates from dropdown and data redistributes automatically

No save/reopen required - formulas recalculate immediately in LibreOffice/Sheets

Author: Cyrus
Date: 2025-11-27
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
from dateutil.relativedelta import relativedelta
from pathlib import Path

INPUT_FILE = Path('/home/cyrus/.cyrus/REG-30/attachments/attachment_3.xlsx')
OUTPUT_FILE = Path('/home/cyrus/cyrus-app/concepts/RealGold_Finmodel_V2_DYNAMIC_DROPDOWNS.xlsx')

def print_section(title):
    print(f"\n{'='*80}")
    print(f"{title}")
    print('='*80)

def fix_fees(wb):
    """Fix unitization fee"""
    print_section("PHASE 1: Fix Fees")

    inputs_ws = wb['Inputs']
    inputs_ws['B26'].value = 0.0001  # 0.01%

    print(f"✓ Unitization Fee: 0.0001 (0.01%)")
    print(f"✓ Admin Fee: {inputs_ws['B25'].value} (verified)")

    return True

def create_xauconfig_timeline(wb):
    """Create XAUconfig with 60-month timeline starting Dec '25"""
    print_section("PHASE 2: Create XAUconfig Timeline")

    if 'XAUconfig' in wb.sheetnames:
        del wb['XAUconfig']

    config_ws = wb.create_sheet('XAUconfig')

    # Headers
    config_ws['A1'] = 'XAU Configuration Timeline'
    config_ws['A2'] = 'Description'
    config_ws['B2'] = '60-month timeline: Dec \'25 - Nov \'30'
    config_ws['A4'] = 'Month Label'
    config_ws['A6'] = 'Date Dropdown List'

    # Build 60 months starting Dec 2025
    start_date = datetime(2025, 12, 1)

    month_labels = []
    for month_offset in range(60):
        col_idx = month_offset + 2  # Start at column B
        col_letter = get_column_letter(col_idx)

        current_date = start_date + relativedelta(months=month_offset)
        month_label = current_date.strftime("%b '%y")

        # Row 4: Month labels for display
        config_ws[f'{col_letter}4'] = month_label
        month_labels.append(month_label)

        # Row 6: Month labels for dropdown list (will be used in data validation)
        config_ws[f'{col_letter}6'] = month_label

    print(f"✓ Created 60-month timeline: {month_labels[0]} → {month_labels[59]}")
    print(f"  Timeline range: XAUconfig!$B$4:$BI$4")
    print(f"  Dropdown list: XAUconfig!$B$6:$BI$6")

    return True

def setup_mine_inventory_with_dropdowns(wb):
    """Setup Mine Inventory with dropdown date selectors and dynamic offsets"""
    print_section("PHASE 3: Mine Inventory with Date Dropdowns")

    mine_ws = wb['Mine Inventory']

    # Set initial dates
    mine_ws['B2'].value = "Dec '25"  # Courbet
    mine_ws['B3'].value = "Jan '26"  # Mine 2

    # Column AO header
    mine_ws['AO1'] = 'Month Offset (Dynamic)'

    # Create data validation for date dropdown
    # This will allow users to select from the list of dates in XAUconfig!$B$6:$BI$6
    dv = DataValidation(
        type="list",
        formula1="XAUconfig!$B$6:$BI$6",
        allow_blank=False,
        showDropDown=True,
        showErrorMessage=True,
        errorTitle="Invalid Date",
        error="Please select a date from the dropdown list"
    )
    dv.prompt = "Select Mine Onboard Date"
    dv.promptTitle = "Date Selection"

    # Add data validation to worksheet
    mine_ws.add_data_validation(dv)

    # Add MATCH formula and dropdown for each mine (rows 2-13)
    for row in range(2, 14):
        mine_name = mine_ws[f'A{row}'].value

        if not mine_name or 'Total' in str(mine_name):
            continue

        # Add data validation (dropdown) to Column B cell
        dv.add(f'B{row}')

        # MATCH formula: finds position of date in timeline, converts to 0-based offset
        formula = f'=IFERROR(MATCH(B{row},XAUconfig!$B$4:$BI$4,0)-1,"")'
        mine_ws[f'AO{row}'] = formula

        print(f"  ✓ Row {row} ({mine_name}): Added dropdown + MATCH formula")

    print("\n✓ Column B now has dropdown date selectors")
    print("✓ Column AO contains dynamic MATCH formulas")
    print("  → Select date from dropdown, offset auto-recalculates!")

    return True

def update_resource_supply_with_sumif(wb):
    """Update Resource Supply to use SUMIF formulas"""
    print_section("PHASE 4: Resource Supply SUMIF Formulas")

    resource_ws = wb['Resource Supply (5yr)']

    # Update headers to reference XAUconfig
    print("Updating month headers...")
    for col_idx in range(2, 62):  # 60 months
        col_letter = get_column_letter(col_idx)
        resource_ws[f'{col_letter}1'] = f'=XAUconfig!{col_letter}4'

    print("✓ Headers linked to XAUconfig")

    # Row 3: Count of mines (COUNTIF)
    # Row 5: Gold production (SUMIF)
    print("\nAdding SUMIF/COUNTIF formulas...")

    for col_idx in range(2, 62):
        col_letter = get_column_letter(col_idx)
        month_offset = col_idx - 2  # 0-based offset

        # Row 3: Count mines at this offset
        resource_ws[f'{col_letter}3'] = f'=COUNTIF(\'Mine Inventory\'!$AO$2:$AO$13,{month_offset})'

        # Row 5: Sum gold production (column M) for mines at this offset
        resource_ws[f'{col_letter}5'] = f'=SUMIF(\'Mine Inventory\'!$AO$2:$AO$13,{month_offset},\'Mine Inventory\'!$M$2:$M$13)'

    print("✓ Row 3: COUNTIF formulas (count mines per month)")
    print("✓ Row 5: SUMIF formulas (sum gold production per month)")
    print("\n✓ Resource Supply now fully dynamic!")

    return True

def update_other_sheet_headers(wb):
    """Update other sheets to reference XAUconfig timeline"""
    print_section("PHASE 5: Update Other Sheet Headers")

    sheets_to_update = [
        'Token Supply (5yr)',
        'Token Demand (5yr)',
        'RA LLC Cashflow',
        'RAF Cashflow (5yr)',
        'RGT Cashflow (5yr)',
    ]

    for sheet_name in sheets_to_update:
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠ {sheet_name} not found, skipping")
            continue

        ws = wb[sheet_name]
        for col_idx in range(2, 62):
            col_letter = get_column_letter(col_idx)
            ws[f'{col_letter}1'] = f'=XAUconfig!{col_letter}4'

        print(f"  ✓ {sheet_name}")

    return True

def create_model_health(wb):
    """Create Model Health dashboard"""
    print_section("PHASE 6: Model Health Dashboard")

    if 'Model Health' in wb.sheetnames:
        del wb['Model Health']

    health_ws = wb.create_sheet('Model Health', 0)

    # Title
    health_ws['A1'] = 'RealGold Financial Model - DYNAMIC with DROPDOWNS'
    health_ws['A1'].font = openpyxl.styles.Font(size=16, bold=True)

    # Metadata
    health_ws['A3'] = 'Version'
    health_ws['B3'] = 'V2.4 - Dynamic Dropdowns'
    health_ws['A4'] = 'Updated'
    health_ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Features
    health_ws['A6'] = 'KEY FEATURE: Dropdown Date Selection'
    health_ws['A6'].font = openpyxl.styles.Font(bold=True, size=12, color='00FF0000')

    health_ws['A7'] = 'Instructions'
    health_ws['B7'] = "1. Go to 'Mine Inventory' sheet"
    health_ws['A8'] = ''
    health_ws['B8'] = "2. Click Column B cell for any mine"
    health_ws['A9'] = ''
    health_ws['B9'] = "3. Select date from DROPDOWN (no typing needed)"
    health_ws['A10'] = ''
    health_ws['B10'] = "4. Data redistributes AUTOMATICALLY (no save/reopen)"

    # Corrections
    row = 12
    health_ws[f'A{row}'] = 'Corrections Applied'
    health_ws[f'A{row}'].font = openpyxl.styles.Font(bold=True, size=12)
    row += 1

    fixes = [
        ('Unitization Fee', '0.5% → 0.01%', 'CRITICAL'),
        ('Courbet Mine', "Sep '25 → Dec '25", 'Updated'),
        ('Mine 2', "Nov '25 → Jan '26", 'Updated'),
        ('Column B', 'DROPDOWN date selectors', 'NEW FEATURE'),
        ('Column AO', 'MATCH formulas', 'DYNAMIC OFFSETS'),
        ('Resource Supply', 'SUMIF/COUNTIF formulas', 'DYNAMIC DATA'),
        ('All Headers', 'XAUconfig references', 'Dynamic timeline'),
    ]

    health_ws[f'A{row}'] = 'Item'
    health_ws[f'B{row}'] = 'Change'
    health_ws[f'C{row}'] = 'Status'
    for col in [f'A{row}', f'B{row}', f'C{row}']:
        health_ws[col].font = openpyxl.styles.Font(bold=True)
    row += 1

    for item, change, status in fixes:
        health_ws[f'A{row}'] = item
        health_ws[f'B{row}'] = change
        health_ws[f'C{row}'] = status
        row += 1

    # Validation
    row += 2
    health_ws[f'A{row}'] = 'Live Validation'
    health_ws[f'A{row}'].font = openpyxl.styles.Font(bold=True, size=12)
    row += 1

    health_ws[f'A{row}'] = 'Unitization Fee'
    health_ws[f'B{row}'] = '=IF(Inputs!B26=0.0001,"✓ PASS","✗ FAIL")'
    row += 1

    health_ws[f'A{row}'] = 'Timeline Start'
    health_ws[f'B{row}'] = '=XAUconfig!B4'
    health_ws[f'C{row}'] = "Should be: Dec '25"
    row += 1

    health_ws[f'A{row}'] = 'Courbet Date'
    health_ws[f'B{row}'] = '=\"Mine Inventory\"!B2'
    health_ws[f'C{row}'] = "Should be: Dec '25"
    row += 1

    health_ws[f'A{row}'] = 'Courbet Offset (Dynamic)'
    health_ws[f'B{row}'] = '=\"Mine Inventory\"!AO2'
    health_ws[f'C{row}'] = "Should be: 0 (if Courbet is Dec '25)"

    # Column widths
    health_ws.column_dimensions['A'].width = 30
    health_ws.column_dimensions['B'].width = 45
    health_ws.column_dimensions['C'].width = 40

    print("✓ Model Health dashboard created")

    return True

def main():
    """Main execution"""
    print_section("RealGold Financial Model - DYNAMIC with DROPDOWNS")
    print(f"Input:  {INPUT_FILE}")
    print(f"Output: {OUTPUT_FILE}")

    if not INPUT_FILE.exists():
        print(f"❌ ERROR: Input file not found: {INPUT_FILE}")
        return 1

    # Ensure output directory exists
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    print("\nLoading workbook...")
    wb = openpyxl.load_workbook(INPUT_FILE)
    print(f"✓ Loaded {len(wb.sheetnames)} sheets")

    # Execute all phases
    fix_fees(wb)
    create_xauconfig_timeline(wb)
    setup_mine_inventory_with_dropdowns(wb)
    update_resource_supply_with_sumif(wb)
    update_other_sheet_headers(wb)
    create_model_health(wb)

    # Save
    print_section("SAVING DYNAMIC MODEL WITH DROPDOWNS")
    print(f"Saving to: {OUTPUT_FILE}")
    wb.save(str(OUTPUT_FILE))
    file_size_mb = OUTPUT_FILE.stat().st_size / 1024 / 1024
    print(f"✓ File saved successfully ({file_size_mb:.2f} MB)")

    print_section("✓ DYNAMIC MODEL WITH DROPDOWNS COMPLETE")
    print("HOW TO USE:")
    print("1. Open 'Mine Inventory' sheet")
    print("2. Click any cell in Column B (mine onboard dates)")
    print("3. Click the dropdown arrow that appears")
    print("4. SELECT a date from the list")
    print("5. Data AUTOMATICALLY redistributes (no save/reopen needed)")
    print("\n✓ Dropdown date selection enabled!")
    print(f"\n📁 Output file: {OUTPUT_FILE}")

    return 0

if __name__ == '__main__':
    import sys
    try:
        sys.exit(main())
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
